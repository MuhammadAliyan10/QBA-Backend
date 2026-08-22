# src/core/extraction/output_formatter.py
"""
OutputFormatter — Enforced Output Format Serialization (Phase 4)

Problem being solved:
  Extracted data is returned as raw Python dict/list.
  The existing CSV export in core_workflow.py is hardcoded and unaware of:
    - CSV column ordering (currently whatever dict.keys() returns)
    - JSON array vs NDJSON vs pretty JSON
    - Required fields validation (non-null enforcement)
    - Partial null fields stripping
    - Excel format (not supported at all)

Solution:
  OutputFormatter takes extracted rows + output format specification and
  returns a formatted bytes payload + MIME type, ready for S3 upload or
  WebSocket streaming.

Supported formats:
  - "csv"    : RFC4180 CSV with ordered columns, CSV injection sanitised
  - "json"   : Compact JSON array
  - "jsonl"  : Newline-delimited JSON (NDJSON) — best for streaming
  - "excel"  : XLSX via openpyxl (optional dependency — degrades to CSV if missing)
  - "tsv"    : Tab-separated values (simpler than CSV for some pipelines)

Schema validation (optional):
  If `required_fields` is provided, rows missing any required field are
  logged as warnings but still included (partial data > no data for MVP).
  A validation_report is returned alongside the data.
"""

from __future__ import annotations

import csv
import io
import json
import logging
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger("output_formatter")

# ---------------------------------------------------------------------------
# CSV injection prevention
# ---------------------------------------------------------------------------
_FORMULA_CHARS = frozenset("=+-@\t\r")


def _sanitize_cell(value: Any) -> Any:
    """Prefix formula-starting strings with ' to prevent CSV injection."""
    if isinstance(value, str) and value and value[0] in _FORMULA_CHARS:
        return f"'{value}"
    return value


def _flatten_row(row: Any) -> Dict[str, Any]:
    """Convert any row shape into a flat dict."""
    if isinstance(row, dict):
        return row
    if isinstance(row, (list, tuple)):
        return {f"col_{i}": v for i, v in enumerate(row)}
    return {"value": str(row)}


# ---------------------------------------------------------------------------
# Validation report
# ---------------------------------------------------------------------------

@dataclass
class ValidationReport:
    total_rows: int = 0
    rows_with_nulls: int = 0
    missing_required: Dict[str, int] = field(default_factory=dict)
    format_used: str = ""

    def to_dict(self) -> dict:
        return {
            "total_rows": self.total_rows,
            "rows_with_nulls": self.rows_with_nulls,
            "missing_required_by_field": self.missing_required,
            "format": self.format_used,
            "completeness_pct": round(
                100 * (1 - self.rows_with_nulls / max(self.total_rows, 1)), 1
            ),
        }


# ---------------------------------------------------------------------------
# Main formatter
# ---------------------------------------------------------------------------

class OutputFormatter:
    """
    Stateless formatter. Call format() with rows and get bytes back.
    """

    SUPPORTED_FORMATS = frozenset({"csv", "json", "jsonl", "tsv", "excel"})

    def format(
        self,
        rows: List[Any],
        *,
        output_format: str = "json",
        column_order: Optional[List[str]] = None,
        required_fields: Optional[List[str]] = None,
        strip_null_fields: bool = False,
    ) -> Tuple[bytes, str, ValidationReport]:
        """
        Serialise extracted rows to the requested format.

        Args:
            rows:            List of dicts (or any flatten-able shape)
            output_format:   One of: csv, json, jsonl, tsv, excel
            column_order:    Explicit column ordering for CSV/TSV/Excel.
                             If None, uses sorted(all_keys) for determinism.
            required_fields: Fields that should be non-null. Violations logged.
            strip_null_fields: Remove keys with null/empty values from each row.

        Returns:
            (payload_bytes, mime_type, validation_report)
        """
        fmt = output_format.lower().strip()
        if fmt not in self.SUPPORTED_FORMATS:
            logger.warning(f"[OutputFormatter] Unknown format '{fmt}', defaulting to json")
            fmt = "json"

        # Normalise rows to list of flat dicts
        flat_rows: List[Dict[str, Any]] = [_flatten_row(r) for r in rows if r is not None]

        if not flat_rows:
            payload = b"[]" if fmt == "json" else b""
            report = ValidationReport(total_rows=0, format_used=fmt)
            return payload, _mime(fmt), report

        # Determine column set and order
        all_keys: List[str] = column_order or sorted(
            {k for row in flat_rows for k in row.keys()}
        )

        # Strip null fields if requested
        if strip_null_fields:
            flat_rows = [
                {k: v for k, v in row.items() if v is not None and v != ""}
                for row in flat_rows
            ]

        # Validate
        report = self._validate(flat_rows, required_fields or [], fmt)

        # Serialise
        try:
            if fmt == "csv":
                payload = self._to_csv(flat_rows, all_keys, delimiter=",")
                return payload, "text/csv; charset=utf-8", report

            elif fmt == "tsv":
                payload = self._to_csv(flat_rows, all_keys, delimiter="\t")
                return payload, "text/tab-separated-values; charset=utf-8", report

            elif fmt == "json":
                payload = json.dumps(flat_rows, ensure_ascii=False, default=str).encode("utf-8")
                return payload, "application/json", report

            elif fmt == "jsonl":
                lines = "\n".join(
                    json.dumps(row, ensure_ascii=False, default=str) for row in flat_rows
                )
                return lines.encode("utf-8"), "application/x-ndjson", report

            elif fmt == "excel":
                payload = self._to_excel(flat_rows, all_keys)
                return payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", report

        except Exception as exc:
            logger.error(f"[OutputFormatter] Serialisation failed for '{fmt}': {exc}. Falling back to JSON.")
            payload = json.dumps(flat_rows, ensure_ascii=False, default=str).encode("utf-8")
            report.format_used = "json (fallback)"
            return payload, "application/json", report

        # Should never reach here
        return b"[]", "application/json", report

    # -----------------------------------------------------------------------
    # Format implementations
    # -----------------------------------------------------------------------

    def _to_csv(self, rows: List[Dict], columns: List[str], delimiter: str = ",") -> bytes:
        buf = io.StringIO()
        writer = csv.DictWriter(
            buf,
            fieldnames=columns,
            delimiter=delimiter,
            extrasaction="ignore",
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\r\n",
        )
        writer.writeheader()
        for row in rows:
            sanitized = {k: _sanitize_cell(row.get(k, "")) for k in columns}
            writer.writerow(sanitized)
        return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility

    def _to_excel(self, rows: List[Dict], columns: List[str]) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.warning("[OutputFormatter] openpyxl not installed. Falling back to CSV.")
            return self._to_csv(rows, columns)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        # Header row with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2D3748")

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(
                15, min(50, len(col_name) + 4)
            )

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row.get(col_name, "")
                ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # -----------------------------------------------------------------------
    # Validation
    # -----------------------------------------------------------------------

    def _validate(
        self,
        rows: List[Dict],
        required_fields: List[str],
        fmt: str,
    ) -> ValidationReport:
        report = ValidationReport(total_rows=len(rows), format_used=fmt)
        missing_counts: Dict[str, int] = {}

        for row in rows:
            has_null = False
            for field_name in required_fields:
                val = row.get(field_name)
                if val is None or val == "":
                    has_null = True
                    missing_counts[field_name] = missing_counts.get(field_name, 0) + 1
            if has_null:
                report.rows_with_nulls += 1

        report.missing_required = missing_counts

        if missing_counts:
            logger.warning(
                f"[OutputFormatter] Validation: {report.rows_with_nulls}/{len(rows)} rows "
                f"have missing required fields: {missing_counts}"
            )

        return report


# ---------------------------------------------------------------------------
# MIME type helper
# ---------------------------------------------------------------------------

def _mime(fmt: str) -> str:
    return {
        "csv":   "text/csv; charset=utf-8",
        "tsv":   "text/tab-separated-values; charset=utf-8",
        "json":  "application/json",
        "jsonl": "application/x-ndjson",
        "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }.get(fmt, "application/octet-stream")


# ---------------------------------------------------------------------------
# Module-level singleton
# ---------------------------------------------------------------------------
_formatter_instance: Optional[OutputFormatter] = None


def get_output_formatter() -> OutputFormatter:
    global _formatter_instance
    if _formatter_instance is None:
        _formatter_instance = OutputFormatter()
    return _formatter_instance
